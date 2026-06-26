"""
migrate_excel.py
----------------
Convierte el Excel original "Control de Inventario" al modelo normalizado
de dos tablas (Productos / Ventas) que usa la app de Luminara.

Uso:
    python migrate_excel.py ruta/al/excel.xlsx

Genera:
    seed_productos.csv
    seed_ventas.csv

Estos archivos se pueden pegar directamente en las pestañas "Productos" y
"Ventas" del Google Sheet (o importarse). El mapeo de canal/estado de pago
es un MEJOR ESFUERZO a partir de las notas del Excel: revisalo antes de
darlo por bueno.
"""
import sys
import re
import unicodedata
import uuid
import csv

import openpyxl

PRODUCTOS_COLS = [
    "sku", "linea", "tono", "categoria", "pedido", "fecha_pedido",
    "costo_unit", "envio_unit", "cantidad_comprada", "precio_venta",
    "activo", "notas",
]
VENTAS_COLS = [
    "venta_id", "fecha", "sku", "cantidad", "canal", "precio_venta",
    "cliente", "metodo_pago", "estado_pago", "notas",
]


def slug(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").upper()
    return text


def split_producto(nombre: str):
    """'06 Hydra Lipgloss' -> ('06', 'Hydra Lipgloss'). 'Travel Brush Set' -> ('', 'Travel Brush Set')."""
    nombre = str(nombre).strip()
    m = re.match(r"^(\d+)\s+(.*)$", nombre)
    if m:
        return m.group(1), m.group(2).strip()
    return "", nombre


def categoria_de(linea: str) -> str:
    l = linea.lower()
    if any(k in l for k in ["lipgloss", "lipliner", "balm", "lip", "labial"]):
        return "Labios"
    if any(k in l for k in ["blush", "highlight", "drops", "rubor", "iluminador"]):
        return "Rostro"
    if any(k in l for k in ["double touch", "touch"]):
        return "Labios y Mejillas"
    if any(k in l for k in ["brush", "brocha", "set"]):
        return "Brochas"
    if any(k in l for k in ["eye", "ojo", "shadow", "liner", "mascara"]):
        return "Ojos"
    return "General"


def num(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    productos = {}
    ventas = []
    pedido_actual = "Pedido 1"

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value

        # Cabecera de pedido
        if isinstance(a, str) and a.strip().upper().startswith("PEDIDO"):
            m = re.search(r"PEDIDO\s*(\d+)", a.upper())
            if m:
                pedido_actual = f"Pedido {m.group(1)}"
            continue

        # Fila de producto: A = cantidad numérica, B = nombre de producto
        if not isinstance(a, (int, float)) or not isinstance(b, str) or not b.strip():
            continue

        cantidad = int(a)
        tono, linea = split_producto(b)
        costo = num(ws.cell(r, 3).value)       # C Costo Unit
        envio = num(ws.cell(r, 4).value)       # D Envío/Unit
        precio_venta = num(ws.cell(r, 9).value)  # I Precio Venta Real
        vendido = int(num(ws.cell(r, 12).value))  # L Vendido
        notas = ws.cell(r, 15).value or ""     # O Notas
        cliente = ws.cell(r, 16).value or ""   # P Cliente
        metodo = ws.cell(r, 17).value or ""    # Q Método de pago
        almacen_vendido = int(num(ws.cell(r, 18).value))  # R Almacén vendido

        base = slug(f"{tono}-{linea}") if tono else slug(linea)
        sku = base
        # Evita colisiones de SKU entre pedidos
        suffix = 1
        while sku in productos and (
            productos[sku]["linea"] != linea or productos[sku]["tono"] != tono
        ):
            suffix += 1
            sku = f"{base}-{suffix}"

        if sku not in productos:
            productos[sku] = {
                "sku": sku,
                "linea": linea,
                "tono": tono,
                "categoria": categoria_de(linea),
                "pedido": pedido_actual,
                "fecha_pedido": "",
                "costo_unit": round(costo, 2),
                "envio_unit": round(envio, 2),
                "cantidad_comprada": cantidad,
                "precio_venta": round(precio_venta, 2),
                "activo": "TRUE",
                "notas": "",
            }
        else:
            productos[sku]["cantidad_comprada"] += cantidad

        # Ventas
        if vendido > 0:
            txt = f"{notas} {metodo}".lower()
            if "apartado" in txt:
                estado = "Apartado"
            elif "falta pago" in txt or "pendiente" in txt:
                estado = "Pendiente"
            else:
                estado = "Pagado"

            if "transfer" in metodo.lower():
                mp = "Transferencia"
            elif "efectivo" in metodo.lower():
                mp = "Efectivo"
            elif "tarjeta" in metodo.lower():
                mp = "Tarjeta"
            else:
                mp = "Otro"

            alm = min(almacen_vendido, vendido)
            directo = vendido - alm

            if alm > 0:
                ventas.append({
                    "venta_id": str(uuid.uuid4())[:8],
                    "fecha": "",
                    "sku": sku,
                    "cantidad": alm,
                    "canal": "Almacén",
                    "precio_venta": round(precio_venta, 2),
                    "cliente": str(cliente).strip(),
                    "metodo_pago": mp,
                    "estado_pago": estado,
                    "notas": str(notas).strip(),
                })
            if directo > 0:
                ventas.append({
                    "venta_id": str(uuid.uuid4())[:8],
                    "fecha": "",
                    "sku": sku,
                    "cantidad": directo,
                    "canal": "Directo",
                    "precio_venta": round(precio_venta, 2),
                    "cliente": str(cliente).strip(),
                    "metodo_pago": mp,
                    "estado_pago": estado,
                    "notas": str(notas).strip(),
                })

    with open("seed_productos.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=PRODUCTOS_COLS)
        w.writeheader()
        for p in productos.values():
            w.writerow(p)

    with open("seed_ventas.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=VENTAS_COLS)
        w.writeheader()
        for v in ventas:
            w.writerow(v)

    print(f"Productos: {len(productos)}  |  Ventas: {len(ventas)}")
    print("Generados: seed_productos.csv, seed_ventas.csv")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "Luminara_Cosmeticos-Control_de_Inventario.xlsx"
    main(src)
